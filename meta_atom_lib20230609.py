'''--------------------------------------------------------
gordon.yiu@gmail.com
Python code for
1. read parameters from excel file
2. interface with Ansys lumerical FDTD solver
3. build the model, add fdtd solver, monitor
4. conduct the simulation. fetch the result transmission and phase
5. export the result to excel files.
6. examples to repeat the tranmssion and phase of meta-atom library
7. save process, progress, informaion in a log files.
8. build a folder and store all simulation fsp and log files
useful code for photonics design of meta-surface.
--------------------------------------------------------'''

import os
import numpy as np
import sys
import time
from datetime import datetime
from os.path import expanduser
sys.path.append("C:\\Program Files\\Lumerical\\v202\\api\\python\\") #Default windows lumapi path
sys.path.append(os.path.dirname(__file__)) #Current directory
import lumapi

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def message_and_log(file_handler, message):
# Function --- message_and_log
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ---very simple python subroutine to print message with timestamp and save to a log file
    from datetime import datetime
    print(datetime.now().strftime('%m/%d %H:%M:%S')+ '.. ' + message)
    file_handler.write(datetime.now().strftime('%m/%d %H:%M:%S')+ '.. ' + message + '\n')

# ++++++++++++++++++++++++++++++++++++++++++++++++++++
def read_parameters(xlsfilename ):
#+++++++++++++++++++++++++++++++++++++++++++++++++++++
#read all parameters and return them as a list
    import openpyxl
    wb = openpyxl.load_workbook(xlsfilename)
    sheet1 = wb['parameters']
    
    firstRow = 6
    firstCol = 1
    nCols = 2  #maximum column number read
    maxRows = 1000  #maximum  rows are read

    allCells = np.array([[cell.value for cell in row] for row in sheet1.iter_rows()])

    # allCells is zero-indexed
    data = allCells[(firstRow-1):(firstRow-1+maxRows),(firstCol-1):(firstCol-1+nCols)]
    
    simulate_param=[]
    for row in data:
        if  row[0] != None:
            simulate_param.append(row.tolist())  #skip 'Y' and send rest to a list
    return simulate_param
   
#+++++++++++++++++++++++++++++++++++++++++++++++++++++
def update_ind_parameters(order_no, log_file, ind_sim_result, \
    inputxlsfilename, outputxlsfilename):
#+++++++++++++++++++++++++++++++++++++++++++++++++++++
#update individual parameter after each small simulation
    import openpyxl
    #get dir name from fiel handler
    dirname= os.path.abspath(log_file.name).split('\\')[-1].split('.')[0]
    
    
    if order_no == 0:
        wb = openpyxl.load_workbook(inputxlsfilename)
        sheet1 = wb['parameters']
        wb.copy_worksheet(sheet1).title= dirname
        wb.save(outputxlsfilename)
    wb = openpyxl.load_workbook(outputxlsfilename)
    sheet2 = wb[dirname]
    

    #update indvisual result to coresponding position
    for column_no, column_content in enumerate(ind_sim_result, start=0):
        sheet2.cell(order_no + 6,column_no+3).value = column_content
    wb.save(outputxlsfilename)
    # wb.save(inputxlsfilename.split('.')[0] + '_result.' + inputxlsfilename.split('.')[1])

def fdtd_solve(log_file, \
    simulate_parameters,\
    input_xslfilename,
    output_xslfilename):
    '''
    --------------------------------------------------------'''
    #####START of fdtd_solve##################
   
    # ---convenient notation of length and time
    um = 1e-6
    nm = 1e-9
    fs = 1e-15

    #True False

    # ----pillar height
    pillar_height= 940 * nm
    lattice_constant= 800 * nm
    # ----python dictionary data-type (key-value pair) to store 'pillar diameter -transmission' result)
    # transmission_of_diameters = {}
    # ----python dictionary data-type (key-value pair) to store 'pillar diameter -phase' result)
    # phase_of_diameters = {}
    
    dirname= os.path.abspath(log_file.name).split('\\')[-1].split('.')[0]

    fdtd = lumapi.FDTD(hide = False)
    
    message_and_log(log_file, 'start to build structure')

    fdtd.switchtolayout()
    fdtd.selectall()
    fdtd.delete()
    # ---add material
    opt_material = fdtd.addmaterial('Dielectric')
    fdtd.setmaterial(opt_material, 'name', 'Amorphous Si')
    fdtd.setmaterial('Amorphous Si', 'Refractive Index', 3.43)

    # ---Glass substrate
    fdtd.addrect()
    fdtd.set('name', 'Substrate-SiO2')
    fdtd.set('material', 'SiO2 (Glass) - Palik')
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('x span', 1.6 * um)
    fdtd.set('y span', 1.6 * um)
    fdtd.set('z max', 0 * um)
    fdtd.set('z min', -5 * um)
    fdtd.set('alpha', 0.5)
    # ---pillar center
    fdtd.addcircle()
    fdtd.set('name', 'pillar center')
    fdtd.set('material', 'Amorphous Si')
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z max', pillar_height)
    fdtd.set('z min', 0 )
    fdtd.set('radius', 200 * nm)

    # ---pillar top-left
    fdtd.addcircle()
    fdtd.set('name', 'pillar top-left')
    fdtd.set('material', 'Amorphous Si')
    fdtd.set('x', -0.5 * lattice_constant)
    fdtd.set('y', 0.5*3**0.5 * lattice_constant)
    fdtd.set('z max', pillar_height)
    fdtd.set('z min', 0)
    fdtd.set('radius',200 * nm)


    # ---pillar top-right
    fdtd.addcircle()
    fdtd.set('name', 'pillar top-right')
    fdtd.set('material', 'Amorphous Si')
    fdtd.set('x', 0.5 * lattice_constant)
    fdtd.set('y', 0.5*3**0.5 * lattice_constant)
    fdtd.set('z max', pillar_height)
    fdtd.set('z min', 0)
    fdtd.set('radius', 200 * nm)

    # ---pillar bottom-left
    fdtd.addcircle()
    fdtd.set('name', 'pillar bottom-left')
    fdtd.set('material', 'Amorphous Si')
    fdtd.set('x', -0.5 * lattice_constant)
    fdtd.set('y', -0.5*3**0.5 * lattice_constant)
    fdtd.set('z max', pillar_height)
    fdtd.set('z min', 0)
    fdtd.set('radius', 200 * nm)


    # ---pillar bottom-right
    fdtd.addcircle()
    fdtd.set('name', 'pillar bottom-right')
    fdtd.set('material', 'Amorphous Si')
    fdtd.set('x', 0.5 * lattice_constant)
    fdtd.set('y', -0.5*3**0.5 * lattice_constant)
    fdtd.set('z max', pillar_height)
    fdtd.set('z min', 0 * um)
    fdtd.set('radius', 200 * nm)


    # ---FDTD----- #
    fdtd.addfdtd()
    fdtd.set('dimension', '3D')
    fdtd.set('simulation time', 5000 * fs)
    #fdtd.set('simulation time', 10 * fs)
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z', 0)
    fdtd.set('x span', lattice_constant)
    fdtd.set('y span', 3**0.5*lattice_constant)
    fdtd.set('z min', -2 * um)
    fdtd.set('z max', +3 * um)
    fdtd.set('index', 1)
    fdtd.set('mesh accuracy', 3)
    fdtd.set('x min bc', 'Periodic')
    fdtd.set('x max bc', 'Periodic')
    fdtd.set('y min bc', 'Periodic')
    fdtd.set('y max bc', 'Periodic')
    fdtd.set('z min bc', 'PML')
    fdtd.set('z max bc', 'PML')
    fdtd.set('pml profile', 2)
    fdtd.set('pml layers', 128)
    fdtd.set('auto shutoff min', 1e-5)

    # ---plane Source----- #
    fdtd.addplane()
    fdtd.set('name', 'Plane source')
    fdtd.set('injection axis', 'z')
    fdtd.set('direction', 'backward')
    fdtd.set('angle theta', 0)
    fdtd.set('phase', 0)
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z', 1 * um)

    #? span for plane source ??==> find it out
    #following 2 lines seems to have no effect, it always sets to 1.6um
    fdtd.set('x span', 1800 * nm)
    fdtd.set('y span', 1800 * nm)
    fdtd.set('center wavelength', 1.55 * um)
    fdtd.set('wavelength span', 0)


    # ---Monitor 1 transmission
    fdtd.addpower()
    fdtd.set('name', 'Transmission at -1um')
    fdtd.set('monitor type', '2D Z-normal')
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z', -1 * um)
    fdtd.set('x span', lattice_constant)
    fdtd.set('y span', 3**0.5*lattice_constant)

    # ---Monitor 2 Pillar top view
    fdtd.addpower()
    fdtd.set('name', 'top view of post')
    fdtd.set('monitor type', '2D Z-normal')
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z', pillar_height/2)
    fdtd.set('x span', lattice_constant)
    fdtd.set('y span', 3**0.5*lattice_constant)

    # ---Monitor 3 Pillar side view
    fdtd.addpower()
    fdtd.set('name', 'side view of post')
    fdtd.set('monitor type', '2D X-normal')
    fdtd.set('x', 0)
    fdtd.set('y', 0)
    fdtd.set('z', 0)
    fdtd.set('y span', 3**0.5*lattice_constant)
    fdtd.set('z span', 2*um)
    message_and_log(log_file, 'save to basic .fsp file')
    dirname= os.path.abspath(log_file.name).split('\\')[-1].split('.')[0]
    message_and_log(log_file, 'save the basic lms file: ')
    file_name = './'+ dirname +'/'+'simulation'
    fdtd.save(file_name) #save the basic fsp file
    No_of_all_simulation = len(simulate_parameters)


    # ---start to loop for every diameter
    for order_no, ind_simulation in enumerate(simulate_parameters):
        diameter = ind_simulation[1]
        message_and_log(log_file, '*****************************************************************')
        message_and_log(log_file, 'simulation #'+str(order_no+1) + ' of '+str(No_of_all_simulation))
        message_and_log(log_file, 'generate fsp for diameter: '+ '{:.3e}'.format(diameter) )

        fdtd.load('simulation')
        #time.sleep(0.3)
        #fdtd.switchtolayout()
        
        fdtd.setnamed('pillar center', 'radius', diameter/2)
        fdtd.setnamed('pillar top-left', 'radius', diameter/2)
        fdtd.setnamed('pillar top-right', 'radius', diameter/2)
        fdtd.setnamed('pillar bottom-left', 'radius', diameter/2)
        fdtd.setnamed('pillar bottom-right', 'radius', diameter/2)

        fdtd.save('simulation_'+str(ind_simulation[0]))

        message_and_log(log_file, 'save to file: '+ 'simulation_'+str(order_no))
       

        transmission = 0
        phase = 0

        fdtd.run()  #if you want to simulate locally just comment fdtd.close above and de-comment this
        
        
        message_and_log(log_file, 'start new fdtd to analyze fsp file ')
        

        transmission = abs(fdtd.transmission('Transmission at -1um'))       
        E_field = fdtd.getdata('Transmission at -1um', 'Ex')
        a = np.reshape(E_field, (E_field.shape[0], E_field.shape[1]))
        phase = fdtd.angle(a)[int(E_field.shape[0]/2)][int(E_field.shape[1]/2)]
        message_and_log(log_file, 'transmission= ' + '{:.3e}'.format(abs(transmission)))
        message_and_log(log_file, 'phase angle= ' + '{:.3e}'.format(phase))
        # result= []
        # result.append([transmission,phase])
        # result.append(transmission)
        message_and_log(log_file, [transmission, phase].__str__())
        update_ind_parameters(order_no, log_file, [transmission, phase], \
        input_xslfilename, output_xslfilename)

    # ---end of loop for every single simulation

    fdtd.close()
    
    #####END  of fdtd_solve ##################



#+++++++++++++++++++++++++++++++++++++++++++++++++++++
if __name__ == '__main__':
    #++++++++++++intialize---------------
    main_dir_name= datetime.now().strftime('%m%d_%H%M%S')
    #----make new dir and change dir to current
    os.chdir(os.path.dirname(__file__))
    os.makedirs('./' + main_dir_name)
    os.chdir(os.path.dirname(__file__) + '/' + main_dir_name)
    log_file= open(main_dir_name +'.log','w+')
    
    #++++++++++++read parameters from xls file+++++++++++++++++
    os.chdir(os.path.dirname(__file__))
    sim_param_filename = 'parameters.xlsx'
    output_filename = 'result.xlsx'
    message_and_log(log_file, 'loading parameters from ' + sim_param_filename)
    sim_parameters = read_parameters(xlsfilename = sim_param_filename)
    message_and_log(log_file, 'Parameters: ')
    message_and_log(log_file, sim_parameters.__str__())
    
    #++++++++++++ simulate ++++++++++++++++++++++++++++++++++++
    message_and_log(log_file, 'simulating.....')
    fdtd_solve(log_file, \
    sim_parameters,\
    sim_param_filename,\
    output_filename)
    
    #close log file
    log_file.close()
    


